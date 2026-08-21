#!/usr/bin/env python3
"""
美团GEO监测 - 统一Excel解析脚本
从Excel报告中提取数据，按产线+平台拆分，输出每个项目一个JSON文件。

用法:
  python3 parse_excel.py <file1.xlsx> [file2.xlsx ...]
  python3 parse_excel.py --dir <directory>   # 解析目录下所有xlsx
  python3 parse_excel.py --merge <json_dir>   # 合并多个Excel到已有JSON

输出: data/{project_key}.json
"""

import os
import sys
import json
import re
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("需要安装 openpyxl: pip3 install openpyxl")
    sys.exit(1)

SCRIPT_DIR = Path(__file__).parent
SITE_DIR = SCRIPT_DIR.parent
DATA_DIR = SITE_DIR / "data"

# 产线 → (显示名, JSON文件key, 问题数, 达标阈值)
PROJECT_MAP = {
    "美团医美-祛痘":   {"label": "美团医美-祛痘",   "key": "qudou",     "qcount": 106, "threshold": 0.80, "emoji": "🔥"},
    "美团医美-美眼":   {"label": "美团医美-美眼",   "key": "meiyan",    "qcount": 115, "threshold": 0.80, "emoji": "👁️"},
    "美团医美-企业版": {"label": "美团医美-企业版", "key": "qiyeban",   "qcount": 5,   "threshold": 0.80, "emoji": "🏢"},
    "家服":            {"label": "美团家服",         "key": "jiafu",     "qcount": 40,  "threshold": 0.80, "emoji": "🏠"},
    "美团-企业版":     {"label": "美团企业版",       "key": "qiyeban",   "qcount": 5,   "threshold": 0.80, "emoji": "🏢"},
    # 茶饮项目 — 产线名可能是"霸王茶姬"/"茶瀑布"/"沪上阿姨"等
    "霸王茶姬":        {"label": "霸王茶姬",         "key": "bawangchaji", "qcount": 100, "threshold": 0.80, "emoji": "🍵"},
    "茶瀑布":          {"label": "茶瀑布",           "key": "chapubu",     "qcount": 100, "threshold": 0.80, "emoji": "🍵"},
    "沪上阿姨":        {"label": "沪上阿姨",         "key": "hushangayi",  "qcount": 100, "threshold": 0.80, "emoji": "🍵"},
}

# 平台名标准化
PLATFORM_ALIAS = {
    "豆包网页版": "豆包",
    "豆包App": "豆包",
    "豆包手机版": "豆包",
    "DeepSeek网页版": "DeepSeek",
    "DeepSeekApp": "DeepSeek",
    "DeepSeek手机版": "DeepSeek",
    "千问网页版": "千问",
    "千问App": "千问",
    "元宝网页版": "元宝",
    "元宝App": "元宝",
}


def parse_filename(filename: str):
    """从文件名提取日期和平台"""
    # 日期: 2026年08月21日 或 20260821
    date_match = re.search(r'(\d{4})年(\d{2})月(\d{2})日', filename)
    if not date_match:
        date_match = re.search(r'(\d{4})(\d{2})(\d{2})', filename)
    if not date_match:
        return None, None

    date_str = f"{date_match.group(1)}-{date_match.group(2)}-{date_match.group(3)}"
    # 短日期 MM/DD
    short_date = f"{date_match.group(2)}/{date_match.group(3)}"

    # 平台
    platform = None
    for key in PLATFORM_ALIAS:
        if key in filename:
            platform = PLATFORM_ALIAS[key]
            break

    return date_str, short_date, platform


def parse_excel(filepath: str) -> list:
    """解析一个Excel文件，返回按(产线,平台)分组的指标列表"""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    # 找关键列索引
    col_map = {}
    for i, h in enumerate(headers):
        if h:
            col_map[str(h).strip()] = i

    required = ['产线', '提及率', '推荐率', '效果统计']
    for r in required:
        if r not in col_map:
            print(f"  ⚠️ 缺少列 '{r}'，跳过 {os.path.basename(filepath)}")
            wb.close()
            return []

    date_str, short_date, platform = parse_filename(os.path.basename(filepath))
    if not platform:
        print(f"  ⚠️ 无法识别平台，跳过 {os.path.basename(filepath)}")
        wb.close()
        return []

    # 按产线分组
    groups = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        chanxian = str(row[col_map['产线']]).strip() if row[col_map['产线']] else None
        if not chanxian:
            continue

        if chanxian not in groups:
            groups[chanxian] = []
        groups[chanxian].append(row)

    results = []
    for chanxian, rows in groups.items():
        total = len(rows)
        if total == 0:
            continue

        # 查找产线配置
        config = PROJECT_MAP.get(chanxian)
        if not config:
            # 自动创建：用产线名作为key
            auto_key = re.sub(r'[^a-zA-Z0-9\u4e00-\u9fff]', '', chanxian)[:20].lower()
            config = {
                "label": chanxian,
                "key": auto_key,
                "qcount": total,
                "threshold": 0.80,
                "emoji": "📊"
            }
            print(f"  ℹ️ 新产线 '{chanxian}' → key={auto_key}")

        # 计算指标
        total_recommend = 0
        total_mention = 0
        has_reco = 0
        has_mention = 0
        no_mention = 0

        for row in rows:
            rec = row[col_map['推荐率']]
            ment = row[col_map['提及率']]
            effect = str(row[col_map['效果统计']] or '')

            total_recommend += float(rec or 0)
            total_mention += float(ment or 0)

            if '推荐' in effect:
                has_reco += 1
            elif '无提及' in effect:
                no_mention += 1
            elif '提及' in effect:
                has_mention += 1
            else:
                no_mention += 1

        avg_recommend = round(total_recommend / total, 4) if total else 0
        avg_mention = round(total_mention / total, 4) if total else 0

        results.append({
            "project_key": config["key"],
            "label": config["label"],
            "chanxian": chanxian,
            "platform": platform,
            "date": date_str,
            "short_date": short_date,
            "total": total,
            "recommend_rate": avg_recommend,
            "mention_rate": avg_mention,
            "has_reco": has_reco,
            "has_mention": has_mention,
            "no_mention": no_mention,
            "threshold": config["threshold"],
            "emoji": config["emoji"],
        })

    wb.close()
    return results


def load_existing(project_key: str) -> dict:
    """加载已有的项目JSON数据"""
    json_path = DATA_DIR / f"{project_key}.json"
    if json_path.exists():
        with open(json_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {
        "project_key": project_key,
        "label": "",
        "threshold": 0.80,
        "emoji": "📊",
        "platforms": {}
    }


def merge_and_save(results: list):
    """将解析结果合并到已有JSON并保存"""
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    # 按 project_key 分组
    by_project = {}
    for r in results:
        pk = r["project_key"]
        if pk not in by_project:
            by_project[pk] = []
        by_project[pk].append(r)

    for pk, items in by_project.items():
        data = load_existing(pk)
        data["label"] = items[0]["label"]
        data["threshold"] = items[0]["threshold"]
        data["emoji"] = items[0]["emoji"]

        for item in items:
            platform = item["platform"]
            if platform not in data["platforms"]:
                data["platforms"][platform] = {
                    "dates": [],
                    "recommend": [],
                    "mention": [],
                    "hasReco": [],
                    "hasMention": [],
                    "noMention": []
                }

            pdat = data["platforms"][platform]

            # 检查日期是否已存在（替换）
            if item["short_date"] in pdat["dates"]:
                idx = pdat["dates"].index(item["short_date"])
                pdat["recommend"][idx] = round(item["recommend_rate"] * 100, 1)
                pdat["mention"][idx] = round(item["mention_rate"] * 100, 1)
                pdat["hasReco"][idx] = item["has_reco"]
                pdat["hasMention"][idx] = item["has_mention"]
                pdat["noMention"][idx] = item["no_mention"]
                print(f"  🔄 更新 {pk}/{platform} {item['short_date']}")
            else:
                pdat["dates"].append(item["short_date"])
                pdat["recommend"].append(round(item["recommend_rate"] * 100, 1))
                pdat["mention"].append(round(item["mention_rate"] * 100, 1))
                pdat["hasReco"].append(item["has_reco"])
                pdat["hasMention"].append(item["has_mention"])
                pdat["noMention"].append(item["no_mention"])
                print(f"  ✅ 新增 {pk}/{platform} {item['short_date']}")

        # 按日期排序
        for platform in data["platforms"]:
            pdat = data["platforms"][platform]
            sorted_indices = sorted(range(len(pdat["dates"])), key=lambda i: pdat["dates"][i])
            pdat["dates"] = [pdat["dates"][i] for i in sorted_indices]
            pdat["recommend"] = [pdat["recommend"][i] for i in sorted_indices]
            pdat["mention"] = [pdat["mention"][i] for i in sorted_indices]
            pdat["hasReco"] = [pdat["hasReco"][i] for i in sorted_indices]
            pdat["hasMention"] = [pdat["hasMention"][i] for i in sorted_indices]
            pdat["noMention"] = [pdat["noMention"][i] for i in sorted_indices]

        # 写入JSON
        json_path = DATA_DIR / f"{pk}.json"
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"  💾 已保存 {json_path.name}")


def main():
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    files = []
    if "--dir" in sys.argv:
        idx = sys.argv.index("--dir")
        if idx + 1 < len(sys.argv):
            d = Path(sys.argv[idx + 1])
            files = sorted(d.glob("*.xlsx"))
            print(f"扫描目录: {d} ({len(files)} 个文件)")
    else:
        for arg in sys.argv[1:]:
            if arg.startswith("--"):
                continue
            p = Path(arg)
            if p.exists():
                files.append(p)
            else:
                print(f"⚠️ 文件不存在: {arg}")

    if not files:
        print("用法: python3 parse_excel.py <file1.xlsx> [file2.xlsx ...]")
        print("      python3 parse_excel.py --dir <directory>")
        return

    all_results = []
    for f in files:
        print(f"\n📄 {f.name}")
        results = parse_excel(str(f))
        all_results.extend(results)

    if all_results:
        print(f"\n📊 共解析 {len(all_results)} 条记录")
        merge_and_save(all_results)
    else:
        print("\n⚠️ 无有效数据")


if __name__ == "__main__":
    main()
